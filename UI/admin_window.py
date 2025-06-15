import datetime
import requests
from PySide6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                               QPushButton, QStackedWidget, QListWidget, QListWidgetItem,
                               QFrame, QSpacerItem, QSizePolicy, QMenuBar, QMenu, QMessageBox,
                               QFormLayout, QDialog, QLineEdit, QDialogButtonBox, QScrollArea,
                               QTableWidget, QTableWidgetItem, QHeaderView, QTextEdit, QTabWidget,
                               QSpinBox, QComboBox, QDateEdit, QCheckBox, QGroupBox, QFileDialog)
from PySide6.QtGui import QPixmap, QColor, QDoubleValidator, QIntValidator
from PySide6.QtCore import Qt, QSize, QDate
from pathlib import Path


class AdminWindow(QMainWindow):
    def __init__(self, token_data):
        super().__init__()
        self.token_data = token_data
        self.access_token = token_data.get('access_token')
        self.setWindowTitle("АИС 'Сифон' - Администратор")
        self.setFixedSize(1440, 900)

        # Базовый URL API
        self.api_url = "http://localhost:8000"

        # Пути к иконкам
        self.icon_dir = Path(__file__).parent / "icons"

        # Основной стек виджетов
        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        # Создаем виджеты для разных разделов
        self.dashboard_widget = self.create_dashboard_widget()
        self.users_widget = self.create_users_widget()
        self.products_widget = self.create_products_widget()
        self.categories_widget = self.create_categories_widget()
        self.orders_widget = self.create_orders_widget()

        self.stack.addWidget(self.dashboard_widget)
        self.stack.addWidget(self.users_widget)
        self.stack.addWidget(self.products_widget)
        self.stack.addWidget(self.categories_widget)
        self.stack.addWidget(self.orders_widget)

        # Создаем навигационное меню
        self.create_navigation()

        # Загружаем начальные данные
        self.load_initial_data()

        # Стилизация
        self.apply_styles()

    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QLabel {
                color: #333;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 12px;
                font-size: 14px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
            QListWidget {
                border: none;
                background-color: transparent;
            }
            QFrame.section-frame {
                background-color: white;
                border-radius: 8px;
                border: 1px solid #ddd;
                padding: 15px;
            }
            .title-label {
                font-size: 20px;
                font-weight: bold;
                color: #2c3e50;
                margin-bottom: 15px;
            }
            .success-button {
                background-color: #2ecc71;
            }
            .success-button:hover {
                background-color: #27ae60;
            }
            .danger-button {
                background-color: #e74c3c;
            }
            .danger-button:hover {
                background-color: #c0392b;
            }
            .warning-button {
                background-color: #f39c12;
            }
            .warning-button:hover {
                background-color: #d35400;
            }
            QTableWidget {
                background-color: white;
                border-radius: 8px;
                border: 1px solid #ddd;
                alternate-background-color: #f9f9f9;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 5px;
                border: none;
            }
            QMenuBar {
                background-color: transparent;
                padding: 5px;
            }
            QMenuBar::item {
                padding: 8px 15px;
                background-color: transparent;
                border-radius: 4px;
            }
            QMenuBar::item:selected {
                background-color: #3498db;
                color: white;
            }
            QMenuBar::item:pressed {
                background-color: #2980b9;
            }
            QTextEdit, QLineEdit, QComboBox, QSpinBox, QDateEdit {
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 5px;
                font-size: 14px;
            }
            QTextEdit:focus, QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDateEdit:focus {
                border: 1px solid #3498db;
            }
            QTabWidget::pane {
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 5px;
                background: white;
            }
            QTabBar::tab {
                padding: 8px 15px;
                background: #ecf0f1;
                border: 1px solid #ddd;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom: 2px solid #3498db;
                font-weight: bold;
            }
        """)

    def get_auth_headers(self):
        """Возвращает заголовки с токеном авторизации"""
        return {
            "Authorization": f"Bearer {self.token_data['access_token']}",
            "Content-Type": "application/json"
        }


    def create_navigation(self):
        """Создает навигационную панель"""
        menubar = self.menuBar()
        menubar.setNativeMenuBar(False)

        # Главное меню
        menu = menubar.addMenu("Меню")

        dashboard_action = menu.addAction("Главная")
        dashboard_action.triggered.connect(lambda: self.stack.setCurrentWidget(self.dashboard_widget))

        users_action = menu.addAction("Пользователи")
        users_action.triggered.connect(lambda: self.stack.setCurrentWidget(self.users_widget))

        products_action = menu.addAction("Товары")
        products_action.triggered.connect(lambda: self.stack.setCurrentWidget(self.products_widget))

        categories_action = menu.addAction("Категории")
        categories_action.triggered.connect(lambda: self.stack.setCurrentWidget(self.categories_widget))

        orders_action = menu.addAction("Заказы")
        orders_action.triggered.connect(lambda: self.stack.setCurrentWidget(self.orders_widget))

        # Кнопка выхода
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        menubar.setCornerWidget(spacer)

        logout_btn = QPushButton("Выход")
        logout_btn.clicked.connect(self.close)
        menubar.setCornerWidget(logout_btn)

    def load_initial_data(self):
        """Загружает начальные данные для админ-панели"""
        self.load_users()
        self.load_products()
        self.load_categories()
        self.load_last_orders()

        # Заглушки для других данных
        self.orders_card.findChild(QLabel, "value").setText("N/A")
        self.revenue_card.findChild(QLabel, "value").setText("N/A")

    def create_dashboard_widget(self):
        """Создает виджет главной панели"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Главная панель администратора")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50;")
        layout.addWidget(title)

        # Статистические карточки
        cards_widget = QWidget()
        cards_layout = QHBoxLayout(cards_widget)
        cards_layout.setContentsMargins(0, 0, 0, 0)
        cards_layout.setSpacing(15)

        # Карточка пользователей
        self.users_card = self.create_stat_card("Пользователи", "0", "#3498db", "users.png")
        cards_layout.addWidget(self.users_card)

        # Карточка товаров
        self.products_card = self.create_stat_card("Товары", "0", "#2ecc71", "products.png")
        cards_layout.addWidget(self.products_card)

        # Карточка заказов
        self.orders_card = self.create_stat_card("Заказы", "0", "#e74c3c", "orders.png")
        cards_layout.addWidget(self.orders_card)

        # Карточка выручки
        self.revenue_card = self.create_stat_card("Выручка", "0 ₽", "#9b59b6", "revenue.png")
        cards_layout.addWidget(self.revenue_card)

        layout.addWidget(cards_widget)

        # Графики и таблицы
        tab_widget = QTabWidget()

        # Вкладка последних заказов
        self.last_orders_table = QTableWidget()
        self.last_orders_table.setColumnCount(5)
        self.last_orders_table.setHorizontalHeaderLabels(["ID", "Клиент", "Сумма", "Статус", "Дата"])
        self.last_orders_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tab_widget.addTab(self.last_orders_table, "Последние заказы")

        # Вкладка популярных товаров
        self.popular_products_table = QTableWidget()
        self.popular_products_table.setColumnCount(4)
        self.popular_products_table.setHorizontalHeaderLabels(["Товар", "Продажи", "Выручка", "Рейтинг"])
        self.popular_products_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tab_widget.addTab(self.popular_products_table, "Популярные товары")

        layout.addWidget(tab_widget)

        return widget

    def load_last_orders(self):
        """Загружает последние заказы за последние 3 дня"""
        try:
            response = requests.get(
                f"{self.api_url}/orders/",
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                orders = response.json()

                # Фильтруем заказы за последние 3 дня
                three_days_ago = datetime.datetime.now() - datetime.timedelta(days=3)
                recent_orders = []
                for order in orders:
                    date_str = order.get("order_date", "")
                    try:
                        order_date = datetime.datetime.fromisoformat(date_str.replace("Z", ""))
                        if order_date >= three_days_ago:
                            recent_orders.append(order)
                    except ValueError:
                        continue

                self.update_last_orders_table(recent_orders)
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить заказы. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def update_last_orders_table(self, orders):
        """Обновляет таблицу последних заказов"""
        self.last_orders_table.setRowCount(len(orders))

        # Устанавливаем количество столбцов и заголовки
        self.last_orders_table.setColumnCount(5)
        self.last_orders_table.setHorizontalHeaderLabels(["ID", "Клиент", "Дата", "Статус", "Сумма"])
        self.last_orders_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        for row, order in enumerate(orders):
            # ID
            id_item = QTableWidgetItem(str(order.get("id", "")))
            id_item.setTextAlignment(Qt.AlignCenter)
            self.last_orders_table.setItem(row, 0, id_item)

            # Клиент (используем customer_id, так как полное имя клиента может не приходить в этом эндпоинте)
            customer_item = QTableWidgetItem(str(order.get("customer_id", "")))
            customer_item.setTextAlignment(Qt.AlignCenter)
            self.last_orders_table.setItem(row, 1, customer_item)

            # Дата
            date_str = order.get("order_date", "")
            try:
                # Парсим дату из формата ISO 8601
                date = datetime.datetime.fromisoformat(date_str.replace("Z", ""))
                date_item = QTableWidgetItem(date.strftime("%d.%m.%Y %H:%M"))
            except ValueError:
                date_item = QTableWidgetItem(date_str)
            date_item.setTextAlignment(Qt.AlignCenter)
            self.last_orders_table.setItem(row, 2, date_item)

            # Статус
            status_item = QTableWidgetItem(order.get("status", ""))
            status_item.setTextAlignment(Qt.AlignCenter)

            # Устанавливаем цвет в зависимости от статуса
            status = order.get("status", "").lower()
            if status == "доставлен":
                status_item.setForeground(QColor("#2ecc71"))  # зеленый
            elif status == "отменен":
                status_item.setForeground(QColor("#e74c3c"))  # красный
            else:
                status_item.setForeground(QColor("#f39c12"))  # оранжевый

            self.last_orders_table.setItem(row, 3, status_item)

            # Сумма
            total_item = QTableWidgetItem(f"{order.get('total_amount', 0):,.2f} ₽".replace(",", " "))
            total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.last_orders_table.setItem(row, 4, total_item)

    def create_stat_card(self, title, value, color, icon_path):
        """Создает карточку статистики с объединённой иконкой и текстом"""
        card = QWidget()
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(20, 20, 20, 20)
        card_layout.setSpacing(10)

        # Иконка + текст вместе
        icon_label = QLabel()
        icon_pixmap = QPixmap(icon_path)
        if not icon_pixmap.isNull():
            icon_pixmap = icon_pixmap.scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            icon_label.setPixmap(icon_pixmap)
        else:
            print(f"[Warning] Icon not found: {icon_path}")
        icon_label.setFixedSize(24, 24)

        text_label = QLabel(title)
        text_label.setStyleSheet("font-size: 14px; font-weight: bold; color: white;")

        title_layout = QHBoxLayout()
        title_layout.setSpacing(8)
        title_layout.setContentsMargins(0, 0, 0, 0)
        title_layout.addWidget(icon_label)
        title_layout.addWidget(text_label)
        title_layout.addStretch()

        title_container = QWidget()
        title_container.setLayout(title_layout)

        # Значение
        value_label = QLabel(value)
        value_label.setObjectName("value")
        value_label.setStyleSheet("font-size: 18px; font-weight: bold; color: white;")

        # Общий стиль карточки
        card.setStyleSheet(f"""
            background-color: {color};
            border-radius: 12px;
        """)

        card_layout.addWidget(title_container)
        card_layout.addWidget(value_label)

        return card

    def create_users_widget(self):
        """Создает виджет управления пользователями"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок и кнопки
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)

        title = QLabel("Управление пользователями")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        refresh_btn = QPushButton("Обновить")
        refresh_btn.clicked.connect(self.load_users)
        header_layout.addWidget(refresh_btn)

        add_user_btn = QPushButton("Добавить пользователя")
        add_user_btn.setStyleSheet("background-color: #2ecc71;")
        add_user_btn.clicked.connect(self.show_add_user_dialog)
        header_layout.addWidget(add_user_btn)

        layout.addWidget(header_widget)

        # Вкладки для разных типов пользователей
        self.tab_widget = QTabWidget()

        # Вкладка клиентов
        self.customers_tab = QWidget()
        self.customers_layout = QVBoxLayout(self.customers_tab)
        self.customers_table = self.create_user_table("Клиенты")
        self.customers_layout.addWidget(self.customers_table)

        # Кнопки действий для клиентов
        self.customers_actions = self.create_user_actions("Клиенты")
        self.customers_layout.addWidget(self.customers_actions)

        self.tab_widget.addTab(self.customers_tab, "Клиенты")

        # Вкладка сотрудников
        self.employees_tab = QWidget()
        self.employees_layout = QVBoxLayout(self.employees_tab)
        self.employees_table = self.create_user_table("Сотрудники")
        self.employees_layout.addWidget(self.employees_table)

        # Кнопки действий для сотрудников
        self.employees_actions = self.create_user_actions("Сотрудники")
        self.employees_layout.addWidget(self.employees_actions)

        self.tab_widget.addTab(self.employees_tab, "Сотрудники")

        # Вкладка поставщиков
        self.suppliers_tab = QWidget()
        self.suppliers_layout = QVBoxLayout(self.suppliers_tab)
        self.suppliers_table = self.create_user_table("Поставщики")
        self.suppliers_layout.addWidget(self.suppliers_table)

        # Кнопки действий для поставщиков
        self.suppliers_actions = self.create_user_actions("Поставщики")
        self.suppliers_layout.addWidget(self.suppliers_actions)

        self.tab_widget.addTab(self.suppliers_tab, "Поставщики")

        layout.addWidget(self.tab_widget)

        return widget

    def create_user_actions(self, user_type):
        """Создает панель действий для пользователей"""
        actions_widget = QWidget()
        actions_layout = QHBoxLayout(actions_widget)
        actions_layout.setContentsMargins(0, 10, 0, 0)

        edit_btn = QPushButton("Изменить")
        edit_btn.setFixedWidth(120)
        edit_btn.setEnabled(False)
        edit_btn.clicked.connect(lambda: self.on_edit_user_clicked(user_type))

        delete_btn = QPushButton("Удалить")
        delete_btn.setFixedWidth(120)
        delete_btn.setEnabled(False)
        delete_btn.setStyleSheet("background-color: #e74c3c;")
        delete_btn.clicked.connect(lambda: self.on_delete_user_clicked(user_type))

        actions_layout.addWidget(edit_btn)
        actions_layout.addWidget(delete_btn)
        actions_layout.addStretch()

        # Сохраняем ссылки на кнопки для последующего использования
        if user_type == "Клиенты":
            self.customers_edit_btn = edit_btn
            self.customers_delete_btn = delete_btn
            self.customers_table.itemSelectionChanged.connect(lambda: self.update_user_actions_buttons("Клиенты"))
        elif user_type == "Сотрудники":
            self.employees_edit_btn = edit_btn
            self.employees_delete_btn = delete_btn
            self.employees_table.itemSelectionChanged.connect(lambda: self.update_user_actions_buttons("Сотрудники"))
        elif user_type == "Поставщики":
            self.suppliers_edit_btn = edit_btn
            self.suppliers_delete_btn = delete_btn
            self.suppliers_table.itemSelectionChanged.connect(lambda: self.update_user_actions_buttons("Поставщики"))

        return actions_widget

    def on_edit_user_clicked(self, user_type):
        """Обработчик нажатия кнопки 'Изменить'"""
        table = self.get_user_table(user_type)
        if not table:
            return

        row = table.currentRow()
        if row >= 0:
            user_id = table.item(row, 0).text()
            users = self.get_user_data(user_type)
            user = next((u for u in users if str(u.get('id', '')) == user_id), None)

            if user:
                self.edit_user(user)

    def on_toggle_user_status_clicked(self, user_type):
        """Обработчик нажатия кнопки 'Блокировать/Разблокировать'"""
        table = self.get_user_table(user_type)
        if not table:
            return

        row = table.currentRow()
        if row >= 0:
            user_id = table.item(row, 0).text()
            users = self.get_user_data(user_type)
            user = next((u for u in users if str(u.get('id', '')) == user_id), None)

            if user:
                new_status = not user.get('is_active', True)
                self.toggle_user_status(user, new_status)

    def on_delete_user_clicked(self, user_type):
        """Обработчик нажатия кнопки 'Удалить'"""
        table = self.get_user_table(user_type)
        if not table:
            return

        row = table.currentRow()
        if row >= 0:
            user_id = table.item(row, 0).text()
            users = self.get_user_data(user_type)
            user = next((u for u in users if str(u.get('id', '')) == user_id), None)

            if user:
                self.delete_user(user)

    def get_user_table(self, user_type):
        """Возвращает таблицу для указанного типа пользователей"""
        if user_type == "Клиенты":
            return self.customers_table
        elif user_type == "Сотрудники":
            return self.employees_table
        elif user_type == "Поставщики":
            return self.suppliers_table
        return None

    def get_user_data(self, user_type):
        """Возвращает данные для указанного типа пользователей"""
        if user_type == "Клиенты":
            return self.customers_data if hasattr(self, 'customers_data') else []
        elif user_type == "Сотрудники":
            return self.employees_data if hasattr(self, 'employees_data') else []
        elif user_type == "Поставщики":
            return self.suppliers_data if hasattr(self, 'suppliers_data') else []
        return []

    def update_user_actions_buttons(self, user_type):
        """Обновляет состояние кнопок действий в зависимости от выбранного пользователя"""
        table = None
        edit_btn = None
        delete_btn = None

        if user_type == "Клиенты":
            table = self.customers_table
            edit_btn = self.customers_edit_btn
            delete_btn = self.customers_delete_btn
        elif user_type == "Сотрудники":
            table = self.employees_table
            edit_btn = self.employees_edit_btn
            delete_btn = self.employees_delete_btn
        elif user_type == "Поставщики":
            table = self.suppliers_table
            edit_btn = self.suppliers_edit_btn
            delete_btn = self.suppliers_delete_btn

        selected_items = table.selectedItems()
        has_selection = len(selected_items) > 0

        edit_btn.setEnabled(has_selection)
        delete_btn.setEnabled(has_selection)

    def load_users(self):
        """Загружает список пользователей"""
        try:
            # Загружаем клиентов
            response = requests.get(
                f"{self.api_url}/customers/",
                headers=self.get_auth_headers()
            )
            if response.status_code == 200:
                customers = response.json()
                self.update_users_table(customers, "Клиенты")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось загрузить клиентов")

            # Загружаем сотрудников
            response = requests.get(
                f"{self.api_url}/employees/",
                headers=self.get_auth_headers()
            )
            if response.status_code == 200:
                employees = response.json()
                self.update_users_table(employees, "Сотрудники")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось загрузить сотрудников")

            # Загружаем поставщиков
            response = requests.get(
                f"{self.api_url}/suppliers/",
                headers=self.get_auth_headers()
            )
            if response.status_code == 200:
                suppliers = response.json()
                self.update_users_table(suppliers, "Поставщики")
                self.suppliers = suppliers  # Сохраняем данные о поставщиках
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось загрузить поставщиков")

            # Обновляем карточку в dashboard
            total_users = len(customers) + len(employees) + len(suppliers)
            self.users_card.findChild(QLabel, "value").setText(str(total_users))

        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def create_user_table(self, user_type):
        """Создает таблицу для определенного типа пользователей"""
        table = QTableWidget()
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.verticalHeader().setVisible(False)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Настраиваем столбцы в зависимости от типа пользователя
        if user_type == "Клиенты":
            table.setColumnCount(5)
            table.setHorizontalHeaderLabels(["ID", "ФИО", "Телефон", "Email", "Адрес"])
            # Настраиваем политику изменения размера столбцов
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)  # ФИО
            table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Телефон
            table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)  # Email
            table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)  # Адрес

        elif user_type == "Сотрудники":
            table.setColumnCount(5)
            table.setHorizontalHeaderLabels(["ID", "ФИО", "Должность", "Телефон", "Дата приема"])
            # Настраиваем политику изменения размера столбцов
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)  # ФИО
            table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)  # Должность
            table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Телефон
            table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Дата приема

        elif user_type == "Поставщики":
            table.setColumnCount(6)
            table.setHorizontalHeaderLabels(["ID", "Название", "Контактное лицо", "Телефон", "Email", "Адрес"])
            # Настраиваем политику изменения размера столбцов
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)  # Название
            table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)  # Контактное лицо
            table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Телефон
            table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)  # Email
            table.horizontalHeader().setSectionResizeMode(5, QHeaderView.Stretch)  # Адрес

        return table

    def update_users_table(self, users, user_type):
        """Обновляет таблицу пользователей"""
        table = None
        if user_type == "Клиенты":
            table = self.customers_table
            # Сортируем клиентов по ID перед отображением
            self.customers_data = sorted(users, key=lambda x: x.get('id', 0))
        elif user_type == "Сотрудники":
            table = self.employees_table
            # Сортируем сотрудников по ID перед отображением
            self.employees_data = sorted(users, key=lambda x: x.get('id', 0))
        elif user_type == "Поставщики":
            table = self.suppliers_table
            # Сортируем поставщиков по ID перед отображением
            self.suppliers_data = sorted(users, key=lambda x: x.get('id', 0))

        if not table:
            return

        table.setRowCount(len(users))
        table.clearSelection()

        # Остальной код метода остается без изменений
        for row, user in enumerate(users):
            if user_type == "Клиенты":
                # ID
                id_item = QTableWidgetItem(str(user.get("id", "")))
                id_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 0, id_item)

                # ФИО
                name_item = QTableWidgetItem(user.get("full_name", ""))
                table.setItem(row, 1, name_item)

                # Телефон
                phone_item = QTableWidgetItem(user.get("phone", ""))
                table.setItem(row, 2, phone_item)

                # Email
                email_item = QTableWidgetItem(user.get("email", ""))
                table.setItem(row, 3, email_item)

                # Адрес
                address_item = QTableWidgetItem(user.get("address", ""))
                table.setItem(row, 4, address_item)

            elif user_type == "Сотрудники":
                # ID
                id_item = QTableWidgetItem(str(user.get("id", "")))
                id_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 0, id_item)

                # ФИО
                name_item = QTableWidgetItem(user.get("full_name", ""))
                table.setItem(row, 1, name_item)

                # Должность
                position_item = QTableWidgetItem(user.get("position", ""))
                table.setItem(row, 2, position_item)

                # Телефон
                phone_item = QTableWidgetItem(user.get("phone", ""))
                table.setItem(row, 3, phone_item)

                # Дата приема
                hire_date_item = QTableWidgetItem(user.get("hire_date", ""))
                hire_date_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 4, hire_date_item)

            elif user_type == "Поставщики":
                # ID
                id_item = QTableWidgetItem(str(user.get("id", "")))
                id_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 0, id_item)

                # Название
                name_item = QTableWidgetItem(user.get("name", ""))
                table.setItem(row, 1, name_item)

                # Контактное лицо
                contact_item = QTableWidgetItem(user.get("contact_person", ""))
                table.setItem(row, 2, contact_item)

                # Телефон
                phone_item = QTableWidgetItem(user.get("phone", ""))
                table.setItem(row, 3, phone_item)

                # Email
                email_item = QTableWidgetItem(user.get("email", ""))
                table.setItem(row, 4, email_item)

                # Адрес
                address_item = QTableWidgetItem(user.get("address", ""))
                table.setItem(row, 5, address_item)

                # Обновляем кнопки действий
                self.update_user_actions_buttons(user_type)

    def add_user_actions(self, table, row, column, user):
        """Добавляет кнопки действий для пользователя"""
        action_widget = QWidget()
        action_layout = QHBoxLayout(action_widget)
        action_layout.setContentsMargins(5, 5, 5, 5)
        action_layout.setSpacing(5)

        edit_btn = QPushButton("Изменить")
        edit_btn.setFixedWidth(80)
        edit_btn.clicked.connect(lambda _, u=user: self.edit_user(u))
        action_layout.addWidget(edit_btn)

        status_btn = QPushButton("Заблокировать" if user.get("is_active", True) else "Разблокировать")
        status_btn.setFixedWidth(100)
        status_btn.setStyleSheet(
            "background-color: #f39c12;" if user.get("is_active", True) else "background-color: #2ecc71;")
        status_btn.clicked.connect(lambda _, u=user: self.toggle_user_status(u, not u.get("is_active", True)))
        action_layout.addWidget(status_btn)

        delete_btn = QPushButton("Удалить")
        delete_btn.setFixedWidth(80)
        delete_btn.setStyleSheet("background-color: #e74c3c;")
        delete_btn.clicked.connect(lambda _, u=user: self.delete_user(u))
        action_layout.addWidget(delete_btn)

        action_layout.addStretch()
        table.setCellWidget(row, column, action_widget)

    def delete_user(self, user):
        """Удаление пользователя (общий метод)"""
        # Определяем тип пользователя по его атрибутам
        if 'position' in user:  # Это сотрудник
            self.delete_employee(user)
        elif 'contact_person' in user:  # Это поставщик
            self.delete_supplier(user)
        else:  # Это клиент
            self.delete_customer(user)

    def delete_customer(self, customer):
        """Удаление клиента"""
        if QMessageBox.question(self, "Подтверждение",
                                f"Вы действительно хотите удалить клиента {customer.get('full_name', '')}?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            try:
                response = requests.delete(
                    f"{self.api_url}/customers/{customer['id']}",
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Клиент успешно удален")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось удалить клиента. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def delete_supplier(self, supplier):
        """Удаление поставщика"""
        if QMessageBox.question(self, "Подтверждение",
                                f"Вы действительно хотите удалить поставщика {supplier.get('name', '')}?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            try:
                response = requests.delete(
                    f"{self.api_url}/suppliers/{supplier['id']}",
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Поставщик успешно удален")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось удалить поставщика. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def delete_employee(self, employee):
        """Удаление сотрудника"""
        if QMessageBox.question(self, "Подтверждение",
                                f"Вы действительно хотите удалить сотрудника {employee.get('full_name', '')}?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            try:
                response = requests.delete(
                    f"{self.api_url}/employees/{employee['id']}",
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Сотрудник успешно удален")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось удалить сотрудника. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def show_add_user_dialog(self):
        """Показывает диалог выбора типа пользователя для добавления"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавление пользователя")
        dialog.setFixedSize(300, 200)

        layout = QVBoxLayout(dialog)

        label = QLabel("Выберите тип пользователя:")
        layout.addWidget(label)

        btn_customer = QPushButton("Клиент")
        btn_customer.clicked.connect(lambda: self.show_add_customer_dialog(dialog))
        layout.addWidget(btn_customer)

        btn_supplier = QPushButton("Поставщик")
        btn_supplier.clicked.connect(lambda: self.show_add_supplier_dialog(dialog))
        layout.addWidget(btn_supplier)

        btn_employee = QPushButton("Сотрудник")
        btn_employee.clicked.connect(lambda: self.show_add_employee_dialog(dialog))
        layout.addWidget(btn_employee)

        dialog.exec_()

    def show_add_customer_dialog(self, parent_dialog=None):
        """Показывает диалог добавления клиента"""
        if parent_dialog:
            parent_dialog.close()

        dialog = QDialog(self)
        dialog.setWindowTitle("Добавление клиента")
        dialog.setFixedSize(400, 450)

        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Поля для клиента
        full_name_edit = QLineEdit()
        phone_edit = QLineEdit()
        email_edit = QLineEdit()
        address_edit = QLineEdit()
        username_edit = QLineEdit()
        password_edit = QLineEdit()
        password_edit.setEchoMode(QLineEdit.Password)

        form_layout.addRow("ФИО:", full_name_edit)
        form_layout.addRow("Телефон:", phone_edit)
        form_layout.addRow("Email:", email_edit)
        form_layout.addRow("Адрес:", address_edit)
        form_layout.addRow("Логин:", username_edit)
        form_layout.addRow("Пароль:", password_edit)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            new_customer = {
                "full_name": full_name_edit.text(),
                "phone": phone_edit.text(),
                "email": email_edit.text(),
                "address": address_edit.text(),
                "username": username_edit.text(),
                "password": password_edit.text()
            }

            try:
                response = requests.post(
                    f"{self.api_url}/customers/",
                    json=new_customer,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 201:
                    QMessageBox.information(self, "Успех", "Клиент успешно добавлен")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось добавить клиента. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def show_add_supplier_dialog(self, parent_dialog=None):
        """Показывает диалог добавления поставщика"""
        if parent_dialog:
            parent_dialog.close()

        dialog = QDialog(self)
        dialog.setWindowTitle("Добавление поставщика")
        dialog.setFixedSize(400, 450)

        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Поля для поставщика
        name_edit = QLineEdit()
        contact_edit = QLineEdit()
        phone_edit = QLineEdit()
        email_edit = QLineEdit()
        address_edit = QLineEdit()
        username_edit = QLineEdit()
        password_edit = QLineEdit()
        password_edit.setEchoMode(QLineEdit.Password)

        form_layout.addRow("Название компании:", name_edit)
        form_layout.addRow("Контактное лицо:", contact_edit)
        form_layout.addRow("Телефон:", phone_edit)
        form_layout.addRow("Email:", email_edit)
        form_layout.addRow("Адрес:", address_edit)
        form_layout.addRow("Логин:", username_edit)
        form_layout.addRow("Пароль:", password_edit)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            new_supplier = {
                "name": name_edit.text(),
                "contact_person": contact_edit.text(),
                "phone": phone_edit.text(),
                "email": email_edit.text(),
                "address": address_edit.text(),
                "username": username_edit.text(),
                "password": password_edit.text()
            }

            try:
                response = requests.post(
                    f"{self.api_url}/suppliers/",
                    json=new_supplier,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 201:
                    QMessageBox.information(self, "Успех", "Поставщик успешно добавлен")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось добавить поставщика. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def show_add_employee_dialog(self, parent_dialog=None):
        """Показывает диалог добавления сотрудника"""
        if parent_dialog:
            parent_dialog.close()

        dialog = QDialog(self)
        dialog.setWindowTitle("Добавление сотрудника")
        dialog.setFixedSize(400, 500)

        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Поля для сотрудника
        full_name_edit = QLineEdit()
        position_edit = QLineEdit()
        phone_edit = QLineEdit()
        hire_date_edit = QDateEdit()
        hire_date_edit.setCalendarPopup(True)
        hire_date_edit.setDate(QDate.currentDate())
        username_edit = QLineEdit()
        password_edit = QLineEdit()
        password_edit.setEchoMode(QLineEdit.Password)

        form_layout.addRow("ФИО:", full_name_edit)
        form_layout.addRow("Должность:", position_edit)
        form_layout.addRow("Телефон:", phone_edit)
        form_layout.addRow("Дата приема:", hire_date_edit)
        form_layout.addRow("Логин:", username_edit)
        form_layout.addRow("Пароль:", password_edit)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            new_employee = {
                "full_name": full_name_edit.text(),
                "position": position_edit.text(),
                "phone": phone_edit.text(),
                "hire_date": hire_date_edit.date().toString("yyyy-MM-dd"),
                "username": username_edit.text(),
                "password": password_edit.text()
            }

            try:
                response = requests.post(
                    f"{self.api_url}/employees/",
                    json=new_employee,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 201:
                    QMessageBox.information(self, "Успех", "Сотрудник успешно добавлен")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось добавить сотрудника. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def edit_user(self, user):
        """Редактирование пользователя"""
        # Определяем тип пользователя по его атрибутам
        if 'position' in user:  # Это сотрудник
            self.edit_employee(user)
        elif 'contact_person' in user:  # Это поставщик
            self.edit_supplier(user)
        else:  # Это клиент
            self.edit_customer(user)

    def edit_customer(self, customer):
        """Редактирование клиента"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактирование клиента")
        dialog.setFixedSize(400, 500)

        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Поля для клиента
        full_name_edit = QLineEdit(customer.get("full_name", ""))
        phone_edit = QLineEdit(customer.get("phone", ""))
        email_edit = QLineEdit(customer.get("email", ""))
        address_edit = QLineEdit(customer.get("address", ""))
        username_edit = QLineEdit(customer.get("username", ""))
        password_edit = QLineEdit()
        password_edit.setPlaceholderText("Оставьте пустым, чтобы не менять")
        password_edit.setEchoMode(QLineEdit.Password)

        form_layout.addRow("ФИО:", full_name_edit)
        form_layout.addRow("Телефон:", phone_edit)
        form_layout.addRow("Email:", email_edit)
        form_layout.addRow("Адрес:", address_edit)
        form_layout.addRow("Логин:", username_edit)
        form_layout.addRow("Новый пароль:", password_edit)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            updated_customer = {
                "full_name": full_name_edit.text(),
                "phone": phone_edit.text(),
                "email": email_edit.text(),
                "address": address_edit.text(),
                "username": username_edit.text()
            }

            if password_edit.text():
                updated_customer["password"] = password_edit.text()

            try:
                response = requests.put(
                    f"{self.api_url}/customers/{customer['id']}",
                    json=updated_customer,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Данные клиента обновлены")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось обновить клиента. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def edit_supplier(self, supplier):
        """Редактирование поставщика"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактирование поставщика")
        dialog.setFixedSize(400, 450)

        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Поля для поставщика
        name_edit = QLineEdit(supplier.get("name", ""))
        contact_edit = QLineEdit(supplier.get("contact_person", ""))
        phone_edit = QLineEdit(supplier.get("phone", ""))
        email_edit = QLineEdit(supplier.get("email", ""))
        address_edit = QLineEdit(supplier.get("address", ""))

        form_layout.addRow("Название:", name_edit)
        form_layout.addRow("Контактное лицо:", contact_edit)
        form_layout.addRow("Телефон:", phone_edit)
        form_layout.addRow("Email:", email_edit)
        form_layout.addRow("Адрес:", address_edit)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            updated_supplier = {
                "name": name_edit.text(),
                "contact_person": contact_edit.text(),
                "phone": phone_edit.text(),
                "email": email_edit.text(),
                "address": address_edit.text()
            }

            try:
                response = requests.put(
                    f"{self.api_url}/suppliers/{supplier['id']}",
                    json=updated_supplier,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Данные поставщика обновлены")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось обновить поставщика. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def edit_employee(self, employee):
        """Редактирование сотрудника"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактирование сотрудника")
        dialog.setFixedSize(400, 500)

        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Поля для сотрудника
        full_name_edit = QLineEdit(employee.get("full_name", ""))
        position_edit = QLineEdit(employee.get("position", ""))
        phone_edit = QLineEdit(employee.get("phone", ""))
        hire_date_edit = QDateEdit()
        hire_date_edit.setCalendarPopup(True)
        try:
            hire_date = QDate.fromString(employee.get("hire_date", ""), "yyyy-MM-dd")
            hire_date_edit.setDate(hire_date)
        except:
            hire_date_edit.setDate(QDate.currentDate())

        username_edit = QLineEdit(employee.get("username", ""))
        password_edit = QLineEdit()
        password_edit.setPlaceholderText("Оставьте пустым, чтобы не менять")
        password_edit.setEchoMode(QLineEdit.Password)

        form_layout.addRow("ФИО:", full_name_edit)
        form_layout.addRow("Должность:", position_edit)
        form_layout.addRow("Телефон:", phone_edit)
        form_layout.addRow("Дата приема:", hire_date_edit)
        form_layout.addRow("Логин:", username_edit)
        form_layout.addRow("Новый пароль:", password_edit)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            updated_employee = {
                "full_name": full_name_edit.text(),
                "position": position_edit.text(),
                "phone": phone_edit.text(),
                "hire_date": hire_date_edit.date().toString("yyyy-MM-dd"),
                "username": username_edit.text()
            }

            if password_edit.text():
                updated_employee["password"] = password_edit.text()

            try:
                response = requests.put(
                    f"{self.api_url}/employees/{employee['id']}",
                    json=updated_employee,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Данные сотрудника обновлены")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось обновить сотрудника. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def toggle_user_status(self, user, is_active):
        """Блокировка/разблокировка пользователя"""
        try:
            response = requests.patch(
                f"{self.api_url}/users/{user['id']}/status",
                json={"is_active": is_active},
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                QMessageBox.information(self, "Успех",
                                        "Пользователь активирован" if is_active else "Пользователь заблокирован")
                self.load_users()
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось изменить статус. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def create_products_widget(self):
        """Создает виджет управления товарами с горизонтальной прокруткой"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок и кнопки
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)

        title = QLabel("Управление товарами")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        refresh_btn = QPushButton("Обновить")
        refresh_btn.clicked.connect(self.load_products)
        header_layout.addWidget(refresh_btn)

        add_product_btn = QPushButton("Добавить товар")
        add_product_btn.setStyleSheet("background-color: #2ecc71;")
        add_product_btn.clicked.connect(self.show_add_product_dialog)
        header_layout.addWidget(add_product_btn)

        layout.addWidget(header_widget)

        # Фильтры
        filter_widget = QWidget()
        filter_layout = QHBoxLayout(filter_widget)
        filter_layout.setContentsMargins(0, 0, 0, 0)
        filter_layout.setSpacing(15)

        # Фильтр по названию
        name_filter_label = QLabel("Название:")
        self.name_filter_edit = QLineEdit()
        self.name_filter_edit.setPlaceholderText("Введите название товара")
        filter_layout.addWidget(name_filter_label)
        filter_layout.addWidget(self.name_filter_edit)

        # Фильтр по категории
        category_filter_label = QLabel("Категория:")
        self.category_filter_combo = QComboBox()
        self.category_filter_combo.addItem("Все категории", "all")
        for category in getattr(self, 'categories', []):
            self.category_filter_combo.addItem(category.get("name", ""), category.get("id", 0))
        filter_layout.addWidget(category_filter_label)
        filter_layout.addWidget(self.category_filter_combo)

        # Фильтр по поставщику
        supplier_filter_label = QLabel("Поставщик:")
        self.supplier_filter_combo = QComboBox()
        self.supplier_filter_combo.addItem("Все поставщики", "all")
        for supplier in getattr(self, 'suppliers', []):
            self.supplier_filter_combo.addItem(supplier.get("name", ""), supplier.get("id", 0))
        filter_layout.addWidget(supplier_filter_label)
        filter_layout.addWidget(self.supplier_filter_combo)

        # Кнопка применения фильтров
        apply_filter_btn = QPushButton("Применить фильтры")
        apply_filter_btn.clicked.connect(self.apply_product_filters)
        filter_layout.addWidget(apply_filter_btn)

        layout.addWidget(filter_widget)

        # Таблица товаров
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(7)
        self.products_table.setHorizontalHeaderLabels(
            ["ID", "Название", "Категория", "Поставщик", "Цена", "На складе", "Описание"])

        # Настройки для горизонтальной прокрутки
        self.products_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.products_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.products_table.setSizeAdjustPolicy(QTableWidget.AdjustToContents)

        # Фиксируем ширину некоторых столбцов
        self.products_table.setColumnWidth(0, 80)  # ID
        self.products_table.setColumnWidth(1, 200)  # Название
        self.products_table.setColumnWidth(2, 120)  # Категория
        self.products_table.setColumnWidth(3, 120)  # Поставщик
        self.products_table.setColumnWidth(4, 100)  # Цена
        self.products_table.setColumnWidth(5, 100)  # На складе
        # Описание оставляем растягивающимся

        # Остальные настройки таблицы
        self.products_table.verticalHeader().setVisible(False)
        self.products_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.products_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.products_table.setSelectionMode(QTableWidget.SingleSelection)
        self.products_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Добавляем таблицу в контейнер
        layout.addWidget(self.products_table)

        # Панель действий
        action_widget = QWidget()
        action_layout = QHBoxLayout(action_widget)
        action_layout.setContentsMargins(0, 10, 0, 0)

        self.edit_product_btn = QPushButton("Изменить")
        self.edit_product_btn.setFixedWidth(120)
        self.edit_product_btn.setEnabled(False)
        self.edit_product_btn.clicked.connect(self.on_edit_product_clicked)
        action_layout.addWidget(self.edit_product_btn)

        self.delete_product_btn = QPushButton("Удалить")
        self.delete_product_btn.setFixedWidth(120)
        self.delete_product_btn.setStyleSheet("background-color: #e74c3c;")
        self.delete_product_btn.setEnabled(False)
        self.delete_product_btn.clicked.connect(self.on_delete_product_clicked)
        action_layout.addWidget(self.delete_product_btn)

        action_layout.addStretch()
        layout.addWidget(action_widget)

        # Подключение сигнала изменения выбора
        self.products_table.itemSelectionChanged.connect(self.on_product_selection_changed)

        return widget

    def apply_product_filters(self):
        """Применяет фильтры к таблице товаров"""
        name_filter = self.name_filter_edit.text().lower()
        category_filter = self.category_filter_combo.currentData()
        supplier_filter = self.supplier_filter_combo.currentData()

        filtered_products = self.products_data

        if name_filter:
            filtered_products = [p for p in filtered_products if name_filter in p.get("name", "").lower()]

        if category_filter != "all":
            filtered_products = [p for p in filtered_products if p.get("category_id") == category_filter]

        if supplier_filter != "all":
            filtered_products = [p for p in filtered_products if p.get("supplier_id") == supplier_filter]

        self.update_products_table(filtered_products)

    def load_products(self):
        """Загружает список товаров с сервера"""
        try:
            response = requests.get(
                f"{self.api_url}/products/",
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                products = response.json()
                self.products_data = products  # Сохраняем данные для последующего использования
                self.update_products_table(products)

                # Обновляем карточку в dashboard
                self.products_card.findChild(QLabel, "value").setText(str(len(products)))
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить товары. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def update_products_table(self, products):
        """Обновляет таблицу товаров"""
        self.products_table.setRowCount(len(products))

        for row, product in enumerate(products):
            # ID
            id_item = QTableWidgetItem(str(product.get("id", "")))
            id_item.setTextAlignment(Qt.AlignCenter)
            self.products_table.setItem(row, 0, id_item)

            # Название
            name_item = QTableWidgetItem(product.get("name", ""))
            self.products_table.setItem(row, 1, name_item)

            # Категория (ID)
            category_item = QTableWidgetItem(str(product.get("category_id", "")))
            category_item.setTextAlignment(Qt.AlignCenter)
            self.products_table.setItem(row, 2, category_item)

            # Поставщик (ID)
            supplier_item = QTableWidgetItem(str(product.get("supplier_id", "")))
            supplier_item.setTextAlignment(Qt.AlignCenter)
            self.products_table.setItem(row, 3, supplier_item)

            # Цена
            price_item = QTableWidgetItem(f"{product.get('price', 0):,.2f} ₽".replace(",", " "))
            price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.products_table.setItem(row, 4, price_item)

            # Количество на складе
            stock_item = QTableWidgetItem(str(product.get("stock_quantity", 0)))
            stock_item.setTextAlignment(Qt.AlignCenter)
            self.products_table.setItem(row, 5, stock_item)

            # Описание
            desc_item = QTableWidgetItem(product.get("description", ""))
            self.products_table.setItem(row, 6, desc_item)

        # После заполнения данных включаем растягивание последнего столбца
        self.products_table.horizontalHeader().setStretchLastSection(True)

    def on_product_selection_changed(self):
        """Обработчик изменения выбранного товара"""
        selected = self.products_table.selectedItems()
        self.edit_product_btn.setEnabled(len(selected) > 0)
        self.delete_product_btn.setEnabled(len(selected) > 0)

    def on_edit_product_clicked(self):
        """Обработчик нажатия кнопки 'Изменить'"""
        selected_row = self.products_table.currentRow()
        if selected_row >= 0:
            product_id = int(self.products_table.item(selected_row, 0).text())
            product = next((p for p in self.products_data if p['id'] == product_id), None)
            if product:
                self.edit_product(product)

    def on_delete_product_clicked(self):
        """Обработчик нажатия кнопки 'Удалить'"""
        selected_row = self.products_table.currentRow()
        if selected_row >= 0:
            product_id = int(self.products_table.item(selected_row, 0).text())
            product = next((p for p in self.products_data if p['id'] == product_id), None)
            if product:
                self.delete_product(product)

    def show_add_product_dialog(self):
        """Показывает диалог добавления товара"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавление товара")
        dialog.setFixedSize(500, 500)

        layout = QVBoxLayout(dialog)

        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Название
        name_edit = QLineEdit()
        name_edit.setPlaceholderText("Название товара")
        form_layout.addRow("Название:", name_edit)

        # Описание
        desc_edit = QTextEdit()
        desc_edit.setPlaceholderText("Описание товара")
        desc_edit.setFixedHeight(100)
        form_layout.addRow("Описание:", desc_edit)

        # Категория
        category_combo = QComboBox()
        # Заполняем категории (предполагается, что они уже загружены)
        for category in getattr(self, 'categories', []):
            category_combo.addItem(category.get("name", ""), category.get("id", 0))
        form_layout.addRow("Категория:", category_combo)

        # Цена
        price_edit = QLineEdit()
        price_edit.setPlaceholderText("Цена")
        price_edit.setValidator(QDoubleValidator(0, 999999, 2))
        form_layout.addRow("Цена:", price_edit)

        # Количество на складе
        stock_edit = QSpinBox()
        stock_edit.setRange(0, 999999)
        stock_edit.setValue(0)
        form_layout.addRow("Количество на складе:", stock_edit)

        # Поставщик
        supplier_combo = QComboBox()
        # Заполняем поставщиков (предполагается, что они уже загружены)
        for supplier in getattr(self, 'suppliers', []):
            supplier_combo.addItem(supplier.get("name", ""), supplier.get("id", 0))
            print("Supplier added:", supplier.get("name", ""), supplier.get("id", 0))  # Отладочный вывод
        form_layout.addRow("Поставщик:", supplier_combo)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            new_product = {
                "name": name_edit.text(),
                "description": desc_edit.toPlainText(),
                "category_id": category_combo.currentData(),
                "price": float(price_edit.text()),
                "stock_quantity": stock_edit.value(),
                "supplier_id": supplier_combo.currentData()
            }

            try:
                response = requests.post(
                    f"{self.api_url}/products/",
                    json=new_product,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 201:
                    QMessageBox.information(self, "Успех", "Товар успешно добавлен")
                    self.load_products()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось добавить товар. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def edit_product(self, product):
        """Редактирование товара и отправка POST-запроса на /stock-operations/ при изменении количества"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактирование товара")
        dialog.setFixedSize(500, 500)

        layout = QVBoxLayout(dialog)

        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Название
        name_edit = QLineEdit(product.get("name", ""))
        name_edit.setPlaceholderText("Название товара")
        form_layout.addRow("Название:", name_edit)

        # Описание
        desc_edit = QTextEdit()
        desc_edit.setPlainText(product.get("description", ""))
        desc_edit.setPlaceholderText("Описание товара")
        desc_edit.setFixedHeight(100)
        form_layout.addRow("Описание:", desc_edit)

        # Категория
        category_combo = QComboBox()
        for category in getattr(self, 'categories', []):
            category_combo.addItem(category.get("name", ""), category.get("id", 0))
            if category.get("id") == product.get("category_id"):
                category_combo.setCurrentIndex(category_combo.count() - 1)
        form_layout.addRow("Категория:", category_combo)

        # Цена
        price_edit = QLineEdit(str(product.get("price", 0)))
        price_edit.setPlaceholderText("Цена")
        price_edit.setValidator(QDoubleValidator(0, 999999, 2))
        form_layout.addRow("Цена:", price_edit)

        # Количество на складе
        stock_edit = QSpinBox()
        stock_edit.setRange(0, 999999)
        stock_edit.setValue(product.get("stock_quantity", 0))
        form_layout.addRow("Количество на складе:", stock_edit)

        # Поставщик
        supplier_combo = QComboBox()
        for supplier in getattr(self, 'suppliers', []):
            supplier_combo.addItem(supplier.get("name", ""), supplier.get("id", 0))
            if supplier.get("id") == product.get("supplier_id"):
                supplier_combo.setCurrentIndex(supplier_combo.count() - 1)
        form_layout.addRow("Поставщик:", supplier_combo)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            updated_product = {
                "name": name_edit.text(),
                "description": desc_edit.toPlainText(),
                "category_id": category_combo.currentData(),
                "price": float(price_edit.text()),
                "stock_quantity": stock_edit.value(),
                "supplier_id": supplier_combo.currentData()
            }

            # Определяем разницу в количестве товара
            old_quantity = product.get("stock_quantity", 0)
            new_quantity = stock_edit.value()
            quantity_diff = new_quantity - old_quantity

            if quantity_diff != 0:
                # Определяем тип операции
                operation_type = "приход" if quantity_diff > 0 else "расход"

                # Получаем корректный employee_id
                print("TOKEN DATA:", self.token_data)
                employee_id = self.token_data.get("employee_id")
                print("EMPLOYEE_ID:", employee_id)

                if employee_id is None or employee_id == 0:
                    QMessageBox.warning(self, "Ошибка", "Некорректный employee_id")
                    return

                # Подготавливаем данные для POST-запроса
                stock_operation_data = {
                    "product_id": product['id'],
                    "operation_type": operation_type,
                    "quantity": abs(quantity_diff),
                    "operation_date": datetime.datetime.now().isoformat(),
                    "employee_id": employee_id
                }

                try:
                    # Отправляем POST-запрос на /stock-operations/
                    response = requests.post(
                        f"{self.api_url}/stock-operations/",
                        json=stock_operation_data,
                        headers=self.get_auth_headers()
                    )

                    if response.status_code == 201:
                        QMessageBox.information(self, "Успех", "Количество товара обновлено и операция записана")
                    else:
                        QMessageBox.warning(self, "Ошибка",
                                            f"Не удалось записать операцию. Код ошибки: {response.status_code}")
                except requests.exceptions.RequestException as e:
                    QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

            # Обновляем данные товара
            try:
                response = requests.put(
                    f"{self.api_url}/products/{product['id']}",
                    json=updated_product,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Товар успешно обновлен")
                    self.load_products()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось обновить товар. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def delete_product(self, product):
        """Удаление товара"""
        if QMessageBox.question(self, "Подтверждение",
                                "Вы действительно хотите удалить этот товар?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            try:
                response = requests.delete(
                    f"{self.api_url}/products/{product['id']}",
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Товар успешно удален")
                    self.load_products()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось удалить товар. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def create_categories_widget(self):
        """Создает виджет управления категориями товаров"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок и кнопки
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)

        title = QLabel("Управление категориями товаров")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        refresh_btn = QPushButton("Обновить")
        refresh_btn.clicked.connect(self.load_categories)
        header_layout.addWidget(refresh_btn)

        add_category_btn = QPushButton("Добавить категорию")
        add_category_btn.setStyleSheet("background-color: #2ecc71;")
        add_category_btn.clicked.connect(self.show_add_category_dialog)
        header_layout.addWidget(add_category_btn)

        layout.addWidget(header_widget)

        # Таблица категорий
        self.categories_table = QTableWidget()
        self.categories_table.setColumnCount(3)  # Уменьшили количество столбцов, так как убрали столбец с действиями
        self.categories_table.setHorizontalHeaderLabels(["ID", "Название", "Родительская категория"])
        self.categories_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.categories_table.verticalHeader().setVisible(False)
        self.categories_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.categories_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.categories_table.setSelectionMode(QTableWidget.SingleSelection)  # Разрешаем выбор только одной строки
        self.categories_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Подключаем сигнал выбора строки
        self.categories_table.itemSelectionChanged.connect(self.on_category_selection_changed)

        layout.addWidget(self.categories_table)

        # Панель действий внизу
        action_widget = QWidget()
        action_layout = QHBoxLayout(action_widget)
        action_layout.setContentsMargins(0, 10, 0, 0)

        # Кнопка "Изменить"
        self.edit_category_btn = QPushButton("Изменить")
        self.edit_category_btn.setFixedWidth(120)
        self.edit_category_btn.setEnabled(False)  # По умолчанию отключена, пока не выбран элемент
        self.edit_category_btn.clicked.connect(self.on_edit_category_clicked)
        action_layout.addWidget(self.edit_category_btn)

        # Кнопка "Удалить"
        self.delete_category_btn = QPushButton("Удалить")
        self.delete_category_btn.setFixedWidth(120)
        self.delete_category_btn.setStyleSheet("background-color: #e74c3c;")
        self.delete_category_btn.setEnabled(False)  # По умолчанию отключена, пока не выбран элемент
        self.delete_category_btn.clicked.connect(self.on_delete_category_clicked)
        action_layout.addWidget(self.delete_category_btn)

        action_layout.addStretch()

        layout.addWidget(action_widget)

        return widget

    def on_category_selection_changed(self):
        """Обработчик изменения выбранной категории"""
        selected_items = self.categories_table.selectedItems()
        if selected_items:
            # Включаем кнопки, если есть выбранная строка
            self.edit_category_btn.setEnabled(True)
            self.delete_category_btn.setEnabled(True)
        else:
            # Отключаем кнопки, если нет выбранной строки
            self.edit_category_btn.setEnabled(False)
            self.delete_category_btn.setEnabled(False)

    def on_edit_category_clicked(self):
        """Обработчик нажатия кнопки 'Изменить'"""
        selected_row = self.categories_table.currentRow()
        if selected_row >= 0:
            # Получаем ID категории из первого столбца выбранной строки
            category_id = int(self.categories_table.item(selected_row, 0).text())

            # Находим категорию в списке
            for category in getattr(self, 'categories', []):
                if category['id'] == category_id:
                    self.edit_category(category)
                    break

    def on_delete_category_clicked(self):
        """Обработчик нажатия кнопки 'Удалить'"""
        selected_row = self.categories_table.currentRow()
        if selected_row >= 0:
            # Получаем ID категории из первого столбца выбранной строки
            category_id = int(self.categories_table.item(selected_row, 0).text())

            # Находим категорию в списке
            for category in getattr(self, 'categories', []):
                if category['id'] == category_id:
                    self.delete_category(category)
                    break



    def load_categories(self):
        """Загружает список категорий"""
        try:
            response = requests.get(
                f"{self.api_url}/product-categories/",
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                categories = response.json()
                self.update_categories_table(categories)
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить категории. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def update_categories_table(self, categories):
        """Обновляет таблицу категорий"""
        self.categories = categories  # Сохраняем список категорий для последующего использования
        self.categories_table.setRowCount(len(categories))

        # Создаем словарь для быстрого поиска категорий по ID
        categories_dict = {cat['id']: cat for cat in categories}

        for row, category in enumerate(categories):
            # ID
            id_item = QTableWidgetItem(str(category.get("id", "")))
            id_item.setTextAlignment(Qt.AlignCenter)
            self.categories_table.setItem(row, 0, id_item)

            # Название
            name_item = QTableWidgetItem(category.get("name", ""))
            self.categories_table.setItem(row, 1, name_item)

            # Родительская категория
            parent_id = category.get("parent_id")
            parent_name = ""
            if parent_id and parent_id in categories_dict:
                parent_name = categories_dict[parent_id].get("name", "")
            parent_item = QTableWidgetItem(parent_name)
            self.categories_table.setItem(row, 2, parent_item)

    def show_add_category_dialog(self):
        """Показывает диалог добавления категории"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавление категории")
        dialog.setFixedSize(400, 250)

        layout = QVBoxLayout(dialog)

        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Название
        name_edit = QLineEdit()
        name_edit.setPlaceholderText("Название категории")
        form_layout.addRow("Название:", name_edit)

        # Родительская категория
        parent_combo = QComboBox()
        parent_combo.addItem("Нет", 0)  # Значение по умолчанию - нет родительской категории

        # Заполняем существующие категории
        for row in range(self.categories_table.rowCount()):
            cat_id = self.categories_table.item(row, 0).text()
            cat_name = self.categories_table.item(row, 1).text()
            parent_combo.addItem(cat_name, int(cat_id))

        form_layout.addRow("Родительская категория:", parent_combo)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            new_category = {
                "name": name_edit.text(),
                "parent_id": parent_combo.currentData()
            }

            try:
                response = requests.post(
                    f"{self.api_url}/product-categories/",
                    json=new_category,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 201:
                    QMessageBox.information(self, "Успех", "Категория успешно добавлена")
                    self.load_categories()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось добавить категорию. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def edit_category(self, category):
        """Редактирование категории"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Редактирование категории")
        dialog.setFixedSize(400, 250)

        layout = QVBoxLayout(dialog)

        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Название
        name_edit = QLineEdit(category.get("name", ""))
        name_edit.setPlaceholderText("Название категории")
        form_layout.addRow("Название:", name_edit)

        # Родительская категория
        parent_combo = QComboBox()
        parent_combo.addItem("Нет", 0)  # Значение по умолчанию - нет родительской категории

        # Заполняем существующие категории, исключая текущую и ее подкатегории
        current_id = category.get("id")
        for row in range(self.categories_table.rowCount()):
            cat_id = int(self.categories_table.item(row, 0).text())
            cat_name = self.categories_table.item(row, 1).text()

            # Проверяем, чтобы не сделать категорию родителем самой себя или ее потомков
            if cat_id != current_id:
                parent_combo.addItem(cat_name, cat_id)

        # Устанавливаем текущее значение родительской категории
        current_parent = category.get("parent_id", 0)
        index = parent_combo.findData(current_parent)
        if index >= 0:
            parent_combo.setCurrentIndex(index)

        form_layout.addRow("Родительская категория:", parent_combo)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            updated_category = {
                "name": name_edit.text(),
                "parent_id": parent_combo.currentData()
            }

            try:
                response = requests.put(
                    f"{self.api_url}/product-categories/{category['id']}",
                    json=updated_category,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Категория успешно обновлена")
                    self.load_categories()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось обновить категорию. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def delete_category(self, category):
        """Удаление категории"""
        if QMessageBox.question(self, "Подтверждение",
                                "Вы действительно хотите удалить эту категорию?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            try:
                response = requests.delete(
                    f"{self.api_url}/product-categories/{category['id']}",
                    headers=self.get_auth_headers()
                )

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Категория успешно удалена")
                    self.load_categories()
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось удалить категорию. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def create_orders_widget(self):
        """Создает виджет управления заказами"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок и кнопки
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)

        title = QLabel("Управление заказами")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        refresh_btn = QPushButton("Обновить")
        refresh_btn.clicked.connect(self.load_orders)
        header_layout.addWidget(refresh_btn)

        layout.addWidget(header_widget)

        # Фильтры
        filter_widget = QWidget()
        filter_layout = QHBoxLayout(filter_widget)
        filter_layout.setContentsMargins(0, 0, 0, 0)
        filter_layout.setSpacing(15)

        # Фильтр по статусу
        status_label = QLabel("Статус:")
        self.status_combo = QComboBox()
        self.status_combo.addItem("Все", "all")
        self.status_combo.addItem("Оплачен", "Оплачен")
        self.status_combo.addItem("Доставлен", "Доставлен")
        self.status_combo.addItem("Завершен", "Завершен")
        self.status_combo.addItem("Отменен", "Отменен")
        filter_layout.addWidget(status_label)
        filter_layout.addWidget(self.status_combo)

        # Фильтр по дате (от)
        date_from_label = QLabel("От:")
        self.date_from_edit = QDateEdit()
        self.date_from_edit.setCalendarPopup(True)
        self.date_from_edit.setDate(QDate.currentDate().addMonths(-1))
        self.date_from_edit.setDisplayFormat("dd.MM.yyyy")
        filter_layout.addWidget(date_from_label)
        filter_layout.addWidget(self.date_from_edit)

        # Фильтр по дате (до)
        date_to_label = QLabel("До:")
        self.date_to_edit = QDateEdit()
        self.date_to_edit.setCalendarPopup(True)
        self.date_to_edit.setDate(QDate.currentDate())
        self.date_to_edit.setDisplayFormat("dd.MM.yyyy")
        filter_layout.addWidget(date_to_label)
        filter_layout.addWidget(self.date_to_edit)

        # Кнопка применения фильтров
        apply_filter_btn = QPushButton("Применить")
        apply_filter_btn.clicked.connect(self.apply_filters)
        filter_layout.addWidget(apply_filter_btn)

        layout.addWidget(filter_widget)

        # Таблица заказов
        self.orders_table = QTableWidget()
        self.orders_table.setColumnCount(5)
        self.orders_table.setHorizontalHeaderLabels(
            ["ID", "Клиент", "Дата", "Сумма", "Статус"])

        # Настройка ширины столбцов
        self.orders_table.setColumnWidth(0, 80)  # ID
        self.orders_table.setColumnWidth(1, 150)  # Клиент
        self.orders_table.setColumnWidth(2, 150)  # Дата
        self.orders_table.setColumnWidth(3, 120)  # Сумма
        self.orders_table.setColumnWidth(4, 120)  # Статус
        self.orders_table.horizontalHeader().setStretchLastSection(True)

        self.orders_table.verticalHeader().setVisible(False)
        self.orders_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.orders_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.orders_table.setSelectionMode(QTableWidget.SingleSelection)
        self.orders_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        layout.addWidget(self.orders_table)

        # Панель действий под таблицей
        self.action_panel = QWidget()
        action_layout = QHBoxLayout(self.action_panel)
        action_layout.setContentsMargins(0, 10, 0, 0)

        # Кнопки действий
        self.details_btn = QPushButton("Просмотреть детали")
        self.details_btn.setFixedWidth(150)
        self.details_btn.setEnabled(False)
        self.details_btn.clicked.connect(self.show_selected_order_details)
        action_layout.addWidget(self.details_btn)

        self.complete_btn = QPushButton("Отметить как доставленный")
        self.complete_btn.setFixedWidth(200)
        self.complete_btn.setEnabled(False)
        self.complete_btn.setStyleSheet("background-color: #2ecc71; color: white;")
        self.complete_btn.clicked.connect(lambda: self.update_selected_order_status("Доставлен"))
        action_layout.addWidget(self.complete_btn)

        self.finish_btn = QPushButton("Завершить заказ")
        self.finish_btn.setFixedWidth(150)
        self.finish_btn.setEnabled(False)
        self.finish_btn.setStyleSheet("background-color: #9b59b6; color: white;")
        self.finish_btn.clicked.connect(lambda: self.update_selected_order_status("Завершен"))
        action_layout.addWidget(self.finish_btn)

        self.cancel_btn = QPushButton("Отменить заказ")
        self.cancel_btn.setFixedWidth(150)
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.setStyleSheet("background-color: #e74c3c; color: white;")
        self.cancel_btn.clicked.connect(lambda: self.update_selected_order_status("Отменен"))
        action_layout.addWidget(self.cancel_btn)

        action_layout.addStretch()
        layout.addWidget(self.action_panel)

        # Инициализация данных
        self.all_orders = []
        self.filtered_orders = []
        self.current_selected_order = None

        # Подключение сигналов
        self.orders_table.itemSelectionChanged.connect(self.update_action_buttons)

        self.load_orders()

        return widget

    def show_selected_order_details(self):
        """Показывает детали выбранного заказа"""
        if self.current_selected_order:
            self.show_order_details(self.current_selected_order)

    def update_selected_order_status(self, new_status):
        """Обновляет статус выбранного заказа"""
        if self.current_selected_order:
            self.update_order_status(self.current_selected_order, new_status)

    def update_action_buttons(self):
        """Обновляет состояние кнопок действий в зависимости от выбранного заказа"""
        selected_items = self.orders_table.selectedItems()
        has_selection = len(selected_items) > 0

        self.details_btn.setEnabled(has_selection)

        if has_selection:
            selected_row = self.orders_table.currentRow()
            order_id = int(self.orders_table.item(selected_row, 0).text())
            self.current_selected_order = next(
                (o for o in self.filtered_orders if o['id'] == order_id), None)

            status = self.current_selected_order.get("status", "Новый")

            self.complete_btn.setEnabled(status == "Оплачен")
            self.finish_btn.setEnabled(status == "Доставлен")
            self.cancel_btn.setEnabled(status in ["Оплачен", "Доставлен"])
        else:
            self.current_selected_order = None
            self.complete_btn.setEnabled(False)
            self.finish_btn.setEnabled(False)
            self.cancel_btn.setEnabled(False)

    def apply_filters(self):
        """Применяет фильтры к загруженным данным"""
        status_filter = self.status_combo.currentData()
        date_from = self.date_from_edit.date()
        date_to = self.date_to_edit.date()

        self.filtered_orders = []

        for order in self.all_orders:
            # Фильтрация по статусу
            if status_filter != "all" and order.get("status") != status_filter:
                continue

            # Фильтрация по дате
            order_date_str = order.get("order_date", "")
            try:
                order_date = QDate.fromString(order_date_str[:10], "yyyy-MM-dd")
                if not (date_from <= order_date <= date_to):
                    continue
            except:
                continue

            self.filtered_orders.append(order)

        self.update_orders_table(self.filtered_orders)
        self.update_orders_stats(self.filtered_orders)

    def load_orders(self):
        """Загружает список всех заказов"""
        try:
            response = requests.get(
                f"{self.api_url}/orders/",
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                self.all_orders = response.json()
                self.apply_filters()  # Применяем фильтры после загрузки
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить заказы. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def update_orders_table(self, orders):
        """Обновляет таблицу заказов"""
        self.orders_table.setRowCount(len(orders))

        status_colors = {
            "Оплачен": "#3498db",  # Синий
            "Доставлен": "#2ecc71",  # Зеленый
            "Завершен": "#9b59b6",  # Фиолетовый
            "Отменен": "#e74c3c",  # Красный
            "Новый": "#f39c12"  # Оранжевый
        }

        for row, order in enumerate(orders):
            # ID
            id_item = QTableWidgetItem(str(order.get("id", "")))
            id_item.setTextAlignment(Qt.AlignCenter)
            self.orders_table.setItem(row, 0, id_item)

            # Клиент (ID + имя если есть)
            customer_id = order.get("customer_id", "")
            customer_name = order.get("customer_name", "")
            customer_text = f"{customer_id}" + (f" ({customer_name})" if customer_name else "")
            customer_item = QTableWidgetItem(customer_text)
            self.orders_table.setItem(row, 1, customer_item)

            # Дата
            date_str = order.get("order_date", "")
            try:
                date = datetime.datetime.fromisoformat(date_str.replace("Z", ""))
                date_item = QTableWidgetItem(date.strftime("%d.%m.%Y %H:%M"))
            except ValueError:
                date_item = QTableWidgetItem(date_str)
            date_item.setTextAlignment(Qt.AlignCenter)
            self.orders_table.setItem(row, 2, date_item)

            # Сумма
            total_item = QTableWidgetItem(f"{order.get('total_amount', 0):,.2f} ₽".replace(",", " "))
            total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.orders_table.setItem(row, 3, total_item)

            # Статус
            status = order.get("status", "Новый")
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)

            # Устанавливаем цвет в зависимости от статуса
            status_item.setForeground(QColor(status_colors.get(status, "#000000")))
            self.orders_table.setItem(row, 4, status_item)

    def update_orders_stats(self, orders):
        """Обновляет статистику по заказам"""
        total_orders = len(orders)
        total_amount = sum(order.get("total_amount", 0) for order in orders)

        # Обновляем карточки в dashboard
        self.orders_card.findChild(QLabel, "value").setText(str(total_orders))
        self.revenue_card.findChild(QLabel, "value").setText(f"{total_amount:,.2f} ₽".replace(",", " "))

    def show_order_details(self, order):
        """Показывает детали заказа с корректным форматированием времени"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Детали заказа #{order.get('id', '')}")
        dialog.setMinimumSize(1100, 600)

        layout = QVBoxLayout(dialog)

        # Функция для форматирования даты
        def format_datetime(dt_str):
            try:
                # Парсим дату из ISO формата (с Z или без)
                dt = datetime.datetime.fromisoformat(dt_str.replace('Z', ''))
                return dt.strftime("%d.%m.%Y %H:%M:%S")
            except ValueError:
                return dt_str  # Возвращаем как есть, если не удалось распарсить

        # Информация о заказе
        info_group = QGroupBox("Информация о заказе")
        info_layout = QFormLayout(info_group)

        order_date = format_datetime(order.get("order_date", ""))

        info_layout.addRow("Номер заказа:", QLabel(str(order.get("id", ""))))
        info_layout.addRow("Клиент ID:", QLabel(str(order.get("customer_id", ""))))
        info_layout.addRow("Дата заказа:", QLabel(order_date))
        info_layout.addRow("Статус:", QLabel(order.get("status", "")))
        info_layout.addRow("Общая сумма:", QLabel(f"{order.get('total_amount', 0):,.2f} ₽".replace(",", " ")))

        layout.addWidget(info_group)

        # Детали заказа
        details_group = QGroupBox("Состав заказа")
        details_layout = QVBoxLayout(details_group)

        # Таблица с товарами
        details_table = QTableWidget()
        details_table.setColumnCount(5)
        details_table.setHorizontalHeaderLabels(["ID", "Товар", "Цена", "Кол-во", "Сумма"])
        details_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        details_table.verticalHeader().setVisible(False)
        details_table.setEditTriggers(QTableWidget.NoEditTriggers)

        try:
            # Загружаем детали заказа - используем order['id'] вместо неопределенного order_id
            response = requests.get(
                f"{self.api_url}/order-details/order/{order['id']}",
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                order_details = response.json()
                details_table.setRowCount(len(order_details))

                for row, detail in enumerate(order_details):
                    # ID товара
                    product_id_item = QTableWidgetItem(str(detail.get("product_id", "")))
                    product_id_item.setTextAlignment(Qt.AlignCenter)
                    details_table.setItem(row, 0, product_id_item)

                    # Название товара
                    product_name = detail.get("product_name", "Неизвестно")
                    product_name_item = QTableWidgetItem(product_name)
                    details_table.setItem(row, 1, product_name_item)

                    # Цена за единицу
                    price_item = QTableWidgetItem(f"{detail.get('price_per_unit', 0):,.2f} ₽".replace(",", " "))
                    price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    details_table.setItem(row, 2, price_item)

                    # Количество
                    quantity_item = QTableWidgetItem(str(detail.get("quantity", 0)))
                    quantity_item.setTextAlignment(Qt.AlignCenter)
                    details_table.setItem(row, 3, quantity_item)

                    # Сумма
                    total_price = detail.get("quantity", 0) * detail.get("price_per_unit", 0)
                    total_item = QTableWidgetItem(f"{total_price:,.2f} ₽".replace(",", " "))
                    total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    details_table.setItem(row, 4, total_item)

                    # Дата добавления (если есть в API)
                    if "order_date" in detail:
                        order_date = format_datetime(detail["order_date"])
                        # Можно добавить дополнительный столбец для даты
            else:
                QMessageBox.warning(dialog, "Ошибка",
                                    f"Не удалось загрузить детали заказа. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(dialog, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

        details_layout.addWidget(details_table)
        layout.addWidget(details_group)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
        button_box.accepted.connect(dialog.accept)
        layout.addWidget(button_box)

        dialog.exec_()

    def update_order_status(self, order, new_status):
        """Обновляет статус заказа через PUT запрос"""
        if QMessageBox.question(self, "Подтверждение",
                                f"Изменить статус заказа #{order['id']} на '{new_status}'?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return

        try:
            # Подготавливаем данные для обновления
            update_data = {
                "customer_id": order.get("customer_id", 0),
                "order_date": order.get("order_date", ""),
                "status": new_status,
                "total_amount": order.get("total_amount", 0)
            }

            response = requests.put(
                f"{self.api_url}/orders/{order['id']}",
                json=update_data,
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                QMessageBox.information(self, "Успех", f"Статус заказа изменен на '{new_status}'")
                self.load_orders()  # Перезагружаем данные
            else:
                error_msg = f"Не удалось изменить статус. Код ошибки: {response.status_code}"
                if response.text:
                    error_msg += f"\nОтвет сервера: {response.text}"
                QMessageBox.warning(self, "Ошибка", error_msg)
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def load_supplies(self):
        """Загружает список поставок"""
        try:
            # Получаем параметры фильтрации
            status_filter = self.supplies_widget.findChild(QComboBox).currentText()
            date_filter = self.supplies_widget.findChild(QDateEdit).date().toString("yyyy-MM-dd")

            params = {}
            if status_filter != "Все":
                params["status"] = status_filter
            if date_filter:
                params["date"] = date_filter

            response = requests.get(
                f"{self.api_url}/supplies/",
                params=params,
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                supplies = response.json()
                self.update_supplies_table(supplies)
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить поставки. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def update_supplies_table(self, supplies):
        """Обновляет таблицу поставок"""
        self.supplies_table.setRowCount(len(supplies))

        for row, supply in enumerate(supplies):
            # ID
            id_item = QTableWidgetItem(str(supply.get("id", "")))
            id_item.setTextAlignment(Qt.AlignCenter)
            self.supplies_table.setItem(row, 0, id_item)

            # Поставщик
            supplier_item = QTableWidgetItem(supply.get("supplier_name", ""))
            self.supplies_table.setItem(row, 1, supplier_item)

            # Дата
            date_str = supply.get("created_at", "")
            try:
                date = datetime.datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%fZ")
                date_item = QTableWidgetItem(date.strftime("%d.%m.%Y %H:%M"))
            except ValueError:
                date_item = QTableWidgetItem(date_str)
            date_item.setTextAlignment(Qt.AlignCenter)
            self.supplies_table.setItem(row, 2, date_item)

            # Статус
            status_item = QTableWidgetItem(supply.get("status", ""))
            status_item.setTextAlignment(Qt.AlignCenter)

            # Цвет статуса
            status = supply.get("status", "").lower()
            if status == "подтверждена":
                status_item.setForeground(QColor("#2ecc71"))  # зеленый
            elif status == "отменена":
                status_item.setForeground(QColor("#e74c3c"))  # красный
            elif status == "выполнена":
                status_item.setForeground(QColor("#3498db"))  # синий
            else:
                status_item.setForeground(QColor("#f39c12"))  # оранжевый

            self.supplies_table.setItem(row, 3, status_item)

            # Товары
            products_text = "\n".join([
                f"{item.get('product_name', '?')} - {item.get('quantity', 0)} шт."
                for item in supply.get("items", [])
            ])
            products_item = QTableWidgetItem(products_text)
            self.supplies_table.setItem(row, 4, products_item)

            # Действия
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(5, 5, 5, 5)
            action_layout.setSpacing(5)

            details_btn = QPushButton("Детали")
            details_btn.setFixedWidth(80)
            details_btn.clicked.connect(lambda _, s=supply: self.show_supply_details(s))
            action_layout.addWidget(details_btn)

            if supply.get("status") == "Ожидает":
                confirm_btn = QPushButton("Подтвердить")
                confirm_btn.setFixedWidth(100)
                confirm_btn.setStyleSheet("background-color: #2ecc71;")
                confirm_btn.clicked.connect(lambda _, s=supply: self.update_supply_status(s, "Подтверждена"))
                action_layout.addWidget(confirm_btn)

                cancel_btn = QPushButton("Отменить")
                cancel_btn.setFixedWidth(80)
                cancel_btn.setStyleSheet("background-color: #e74c3c;")
                cancel_btn.clicked.connect(lambda _, s=supply: self.update_supply_status(s, "Отменена"))
                action_layout.addWidget(cancel_btn)

            if supply.get("status") == "Подтверждена":
                complete_btn = QPushButton("Выполнить")
                complete_btn.setFixedWidth(90)
                complete_btn.setStyleSheet("background-color: #3498db;")
                complete_btn.clicked.connect(lambda _, s=supply: self.update_supply_status(s, "Выполнена"))
                action_layout.addWidget(complete_btn)

            action_layout.addStretch()
            self.supplies_table.setCellWidget(row, 5, action_widget)

    def show_supply_details(self, supply):
        """Показывает детали поставки"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Детали поставки #{supply.get('id', '')}")
        dialog.setFixedSize(600, 500)

        layout = QVBoxLayout(dialog)

        # Информация о поставке
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 8px;
                border: 1px solid #ddd;
                padding: 15px;
            }
            QLabel {
                font-size: 14px;
            }
            QLabel#title {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
            }
        """)
        info_layout = QFormLayout(info_frame)
        info_layout.setRowWrapPolicy(QFormLayout.WrapAllRows)

        info_layout.addRow(QLabel("<b>Информация о поставке</b>"))
        info_layout.addRow("Номер поставки:", QLabel(str(supply.get("id", ""))))
        info_layout.addRow("Дата:", QLabel(supply.get("created_at", "")))
        info_layout.addRow("Статус:", QLabel(supply.get("status", "")))
        info_layout.addRow("Поставщик:", QLabel(supply.get("supplier_name", "")))
        info_layout.addRow("Комментарий:", QLabel(supply.get("comment", "")))

        layout.addWidget(info_frame)

        # Товары в поставке
        products_label = QLabel("<b>Товары в поставке</b>")
        layout.addWidget(products_label)

        products_table = QTableWidget()
        products_table.setColumnCount(4)
        products_table.setHorizontalHeaderLabels(["Товар", "Цена", "Количество", "Сумма"])
        products_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        products_table.verticalHeader().setVisible(False)
        products_table.setEditTriggers(QTableWidget.NoEditTriggers)

        products = supply.get("items", [])
        products_table.setRowCount(len(products))

        for row, product in enumerate(products):
            # Товар
            name_item = QTableWidgetItem(product.get("product_name", ""))
            products_table.setItem(row, 0, name_item)

            # Цена
            price_item = QTableWidgetItem(f"{product.get('price_per_unit', 0):,.2f} ₽".replace(",", " "))
            price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            products_table.setItem(row, 1, price_item)

            # Количество
            quantity_item = QTableWidgetItem(str(product.get("quantity", 0)))
            quantity_item.setTextAlignment(Qt.AlignCenter)
            products_table.setItem(row, 2, quantity_item)

            # Сумма
            total_item = QTableWidgetItem(f"{product.get('total_price', 0):,.2f} ₽".replace(",", " "))
            total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            products_table.setItem(row, 3, total_item)

        layout.addWidget(products_table)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
        button_box.accepted.connect(dialog.accept)
        layout.addWidget(button_box)

        dialog.exec()

    def update_supply_status(self, supply, new_status):
        """Обновляет статус поставки"""
        try:
            response = requests.patch(
                f"{self.api_url}/supplies/{supply['id']}/status",
                json={"status": new_status},
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                QMessageBox.information(self, "Успех", f"Статус поставки изменен на '{new_status}'")
                self.load_supplies()

                # Если поставка выполнена, обновляем данные склада
                if new_status == "Выполнена":
                    self.load_warehouse_data()
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось изменить статус. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def create_warehouse_widget(self):
        """Создает виджет управления складом"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Управление складом")
        title.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
        """)
        layout.addWidget(title)

        # Таблица складских операций
        self.warehouse_table = QTableWidget()
        self.warehouse_table.setColumnCount(5)
        self.warehouse_table.setHorizontalHeaderLabels(
            ["ID", "Товар", "Тип операции", "Количество", "Дата операции"])

        self.warehouse_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.warehouse_table.verticalHeader().setVisible(False)
        self.warehouse_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.warehouse_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.warehouse_table.setSelectionMode(QTableWidget.SingleSelection)
        self.warehouse_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        layout.addWidget(self.warehouse_table)

        return widget

    def show_dummy_warehouse_data(self):
        """Показывает демо-данные склада"""
        QMessageBox.information(
            self,
            "Демо-данные склада",
            "Здесь будут отображаться:\n\n"
            "- Текущие остатки товаров\n"
            "- История движений\n"
            "- Остатки по категориям\n"
            "- Учет поставок\n\n"
            "Функционал в разработке",
            QMessageBox.Ok
        )

    def load_warehouse_data(self):
        """Загружает данные о складских операциях"""
        try:
            response = requests.get(
                f"{self.api_url}/stock-operations/",
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                stock_operations = response.json()
                self.update_warehouse_table(stock_operations)
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить складские операции. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def update_warehouse_table(self, stock_operations):
        """Обновляет таблицу складских операций"""
        self.warehouse_table.setRowCount(len(stock_operations))

        for row, operation in enumerate(stock_operations):
            # ID
            id_item = QTableWidgetItem(str(operation.get("id", "")))
            id_item.setTextAlignment(Qt.AlignCenter)
            self.warehouse_table.setItem(row, 0, id_item)

            # Товар
            product_item = QTableWidgetItem(str(operation.get("product_id", "")))
            product_item.setTextAlignment(Qt.AlignCenter)
            self.warehouse_table.setItem(row, 1, product_item)

            # Тип операции
            operation_type_item = QTableWidgetItem(operation.get("operation_type", ""))
            self.warehouse_table.setItem(row, 2, operation_type_item)

            # Количество
            quantity_item = QTableWidgetItem(str(operation.get("quantity", 0)))
            quantity_item.setTextAlignment(Qt.AlignCenter)
            self.warehouse_table.setItem(row, 3, quantity_item)

            # Дата операции
            date_str = operation.get("operation_date", "")
            try:
                date = datetime.datetime.fromisoformat(date_str.replace("Z", ""))
                date_item = QTableWidgetItem(date.strftime("%d.%m.%Y %H:%M"))
            except ValueError:
                date_item = QTableWidgetItem(date_str)
            date_item.setTextAlignment(Qt.AlignCenter)
            self.warehouse_table.setItem(row, 4, date_item)

    def update_stock_quantity(self, item):
        """Обновляет количество товара на складе и отправляет POST-запрос на /stock-operations/"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Изменение количества товара: {item.get('name', '')}")
        dialog.setFixedSize(300, 200)

        layout = QVBoxLayout(dialog)

        # Текущее количество
        current_label = QLabel(f"Текущее количество: {item.get('stock_quantity', 0)}")
        current_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(current_label)

        # Новое количество
        quantity_edit = QSpinBox()
        quantity_edit.setRange(0, 999999)
        quantity_edit.setValue(item.get('stock_quantity', 0))
        layout.addWidget(quantity_edit)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        if dialog.exec() == QDialog.Accepted:
            new_quantity = quantity_edit.value()
            old_quantity = item.get('stock_quantity', 0)

            # Определяем тип операции
            if new_quantity > old_quantity:
                operation_type = "приход"
                quantity = new_quantity - old_quantity
            elif new_quantity < old_quantity:
                operation_type = "расход"
                quantity = old_quantity - new_quantity
            else:
                return  # Количество не изменилось

            # Подготавливаем данные для POST-запроса
            stock_operation_data = {
                "product_id": item['id'],
                "operation_type": operation_type,
                "quantity": quantity,
                "operation_date": datetime.datetime.now().isoformat(),
                "employee_id": self.token_data.get('user_id', 0)  # ID текущего сотрудника
            }

            try:
                # Отправляем POST-запрос на /stock-operations/
                response = requests.post(
                    f"{self.api_url}/stock-operations/",
                    json=stock_operation_data,
                    headers=self.get_auth_headers()
                )

                if response.status_code == 201:
                    QMessageBox.information(self, "Успех", "Количество товара обновлено и операция записана")
                    self.load_products()  # Обновляем список товаров
                    self.load_warehouse_data()  # Обновляем складские операции
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось обновить количество. Код ошибки: {response.status_code}")
            except requests.exceptions.RequestException as e:
                QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def create_finance_widget(self):
        """Создает виджет финансовых отчетов"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Финансовые отчеты")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50;")
        layout.addWidget(title)

        # Кнопка для создания нового отчета
        generate_report_btn = QPushButton("Создать новый отчет")
        generate_report_btn.setStyleSheet("background-color: #2ecc71;")
        generate_report_btn.clicked.connect(self.show_generate_report_dialog)
        layout.addWidget(generate_report_btn)

        # Таблица финансовых отчетов
        self.finance_table = QTableWidget()
        self.finance_table.setColumnCount(5)
        self.finance_table.setHorizontalHeaderLabels(["ID", "Дата отчета", "Общий доход", "Общие расходы", "Прибыль"])
        self.finance_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.finance_table.verticalHeader().setVisible(False)
        self.finance_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.finance_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.finance_table.setSelectionMode(QTableWidget.SingleSelection)
        self.finance_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        layout.addWidget(self.finance_table)

        # Кнопки для экспорта
        export_widget = QWidget()
        export_layout = QHBoxLayout(export_widget)
        export_layout.setContentsMargins(0, 10, 0, 0)

        export_to_excel_btn = QPushButton("Экспорт в Excel")
        export_to_excel_btn.setStyleSheet("background-color: #2ecc71;")
        export_to_excel_btn.clicked.connect(self.export_to_excel)
        export_layout.addWidget(export_to_excel_btn)

        export_to_pdf_btn = QPushButton("Экспорт в PDF")
        export_to_pdf_btn.setStyleSheet("background-color: #e74c3c;")
        export_to_pdf_btn.clicked.connect(self.export_to_pdf)
        export_layout.addWidget(export_to_pdf_btn)

        layout.addWidget(export_widget)

        # Загрузка данных
        self.load_finance_reports()

        return widget

    def show_generate_report_dialog(self):
        """Показывает диалоговое окно для создания нового отчета"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Создать новый отчет")
        dialog.setFixedSize(400, 200)

        layout = QVBoxLayout(dialog)

        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Выбор диапазона дат
        date_from_label = QLabel("От:")
        self.date_from_edit = QDateEdit()
        self.date_from_edit.setCalendarPopup(True)
        self.date_from_edit.setDate(QDate.currentDate().addMonths(-1))
        self.date_from_edit.setDisplayFormat("dd.MM.yyyy")

        date_to_label = QLabel("До:")
        self.date_to_edit = QDateEdit()
        self.date_to_edit.setCalendarPopup(True)
        self.date_to_edit.setDate(QDate.currentDate())
        self.date_to_edit.setDisplayFormat("dd.MM.yyyy")

        form_layout.addRow(date_from_label, self.date_from_edit)
        form_layout.addRow(date_to_label, self.date_to_edit)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(lambda: self.generate_report(dialog))
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.exec_()

    def generate_report(self, dialog):
        """Генерирует новый отчет на основе выбранных параметров"""
        date_from = self.date_from_edit.date().toString("yyyy-MM-dd")
        date_to = self.date_to_edit.date().toString("yyyy-MM-dd")

        try:
            # Получаем данные о заказах за выбранный период
            response = requests.get(
                f"{self.api_url}/orders/",
                params={"start_date": date_from, "end_date": date_to},
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                orders = response.json()
                total_revenue = sum(order.get("total_amount", 0) for order in orders)

                # Получаем данные о расходах за выбранный период
                expenses_response = requests.get(
                    f"{self.api_url}/expenses/",
                    params={"start_date": date_from, "end_date": date_to},
                    headers=self.get_auth_headers()
                )

                if expenses_response.status_code == 200:
                    expenses = expenses_response.json()
                    total_expenses = sum(expense.get("amount", 0) for expense in expenses)

                    # Рассчитываем прибыль
                    profit = total_revenue - total_expenses

                    # Создаем новый отчет
                    new_report = {
                        "report_date": QDate.currentDate().toString("yyyy-MM-dd"),
                        "total_revenue": total_revenue,
                        "total_expenses": total_expenses,
                        "profit": profit
                    }

                    # Отправляем отчет на сервер
                    response = requests.post(
                        f"{self.api_url}/reports/",
                        json=new_report,
                        headers=self.get_auth_headers()
                    )

                    if response.status_code == 201:
                        QMessageBox.information(self, "Успех", "Новый отчет успешно создан")
                        self.load_finance_reports()
                    else:
                        QMessageBox.warning(self, "Ошибка",
                                            f"Не удалось создать отчет. Код ошибки: {response.status_code}")
                else:
                    QMessageBox.warning(self, "Ошибка",
                                        f"Не удалось загрузить расходы. Код ошибки: {expenses_response.status_code}")
            else:
                QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить заказы. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

        dialog.accept()

    def load_finance_reports(self):
        """Загружает финансовые отчеты"""
        try:
            response = requests.get(
                f"{self.api_url}/reports/",
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                reports = response.json()
                self.update_finance_reports_table(reports)
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить финансовые отчеты. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def update_finance_reports_table(self, reports):
        """Обновляет таблицу финансовых отчетов"""
        self.finance_table.setRowCount(len(reports))

        for row, report in enumerate(reports):
            # ID
            id_item = QTableWidgetItem(str(report.get("id", "")))
            id_item.setTextAlignment(Qt.AlignCenter)
            self.finance_table.setItem(row, 0, id_item)

            # Дата отчета
            date_item = QTableWidgetItem(report.get("report_date", ""))
            date_item.setTextAlignment(Qt.AlignCenter)
            self.finance_table.setItem(row, 1, date_item)

            # Общий доход
            revenue_item = QTableWidgetItem(f"{report.get('total_revenue', 0):,.2f} ₽".replace(",", " "))
            revenue_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.finance_table.setItem(row, 2, revenue_item)

            # Общие расходы
            expenses_item = QTableWidgetItem(f"{report.get('total_expenses', 0):,.2f} ₽".replace(",", " "))
            expenses_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.finance_table.setItem(row, 3, expenses_item)

            # Прибыль
            profit_item = QTableWidgetItem(f"{report.get('profit', 0):,.2f} ₽".replace(",", " "))
            profit_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.finance_table.setItem(row, 4, profit_item)

    def export_to_excel(self):
        """Экспорт финансовых отчетов в Excel"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Финансовые отчеты"

        # Заголовки
        headers = ["ID", "Дата отчета", "Общий доход", "Общие расходы", "Прибыль"]
        ws.append(headers)

        # Данные
        for row in range(self.finance_table.rowCount()):
            row_data = []
            for col in range(self.finance_table.columnCount()):
                item = self.finance_table.item(row, col)
                row_data.append(item.text() if item else "")
            ws.append(row_data)

        # Сохранение файла
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчет", "", "Excel Files (*.xlsx)")
        if file_path:
            wb.save(file_path)
            QMessageBox.information(self, "Успех", "Отчет успешно сохранен в Excel")

    def export_to_pdf(self):
        """Экспорт финансовых отчетов в PDF"""
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Заголовки
        headers = ["ID", "Дата отчета", "Общий доход", "Общие расходы", "Прибыль"]
        col_width = pdf.w / 5.5  # Ширина колонки
        for header in headers:
            pdf.cell(col_width, 10, header, border=1)
        pdf.ln()

        # Данные
        for row in range(self.finance_table.rowCount()):
            for col in range(self.finance_table.columnCount()):
                item = self.finance_table.item(row, col)
                pdf.cell(col_width, 10, item.text() if item else "", border=1)
            pdf.ln()

        # Сохранение файла
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчет", "", "PDF Files (*.pdf)")
        if file_path:
            pdf.output(file_path)
            QMessageBox.information(self, "Успех", "Отчет успешно сохранен в PDF")

    def create_finance_stat_card(self, title, value, color, icon_name):
        """Создает карточку финансовой статистики"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border-radius: 8px;
                border: 1px solid #ddd;
                padding: 15px;
            }}
            QLabel#title {{
                font-size: 14px;
                color: #7f8c8d;
                margin-bottom: 5px;
            }}
            QLabel#value {{
                font-size: 18px;
                font-weight: bold;
                color: {color};
                margin-top: 5px;
            }}
        """)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(5)

        # Заголовок и иконка
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)

        title_label = QLabel(title)
        title_label.setObjectName("title")
        header_layout.addWidget(title_label)

        icon_label = QLabel()
        icon_label.setPixmap(QPixmap(str(self.icon_dir / icon_name)).scaled(20, 20, Qt.KeepAspectRatio))
        header_layout.addWidget(icon_label)

        layout.addWidget(header_widget)

        # Значение
        value_label = QLabel(value)
        value_label.setObjectName("value")
        value_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(value_label)

        return card

    def load_finance_data(self):
        """Загружает финансовые данные"""
        try:
            response = requests.get(
                f"{self.api_url}/finance/",
                headers=self.get_auth_headers()
            )

            if response.status_code == 200:
                finance_data = response.json()
                self.update_finance_table(finance_data)
            else:
                QMessageBox.warning(self, "Ошибка",
                                    f"Не удалось загрузить финансовые данные. Код ошибки: {response.status_code}")
        except requests.exceptions.RequestException as e:
            QMessageBox.critical(self, "Ошибка подключения", f"Не удалось подключиться к серверу: {str(e)}")

    def update_finance_table(self, finance_data):
        """Обновляет таблицу финансовых данных"""
        self.finance_table.setRowCount(len(finance_data))

        for row, data in enumerate(finance_data):
            # Дата
            date_item = QTableWidgetItem(data.get("date", ""))
            self.finance_table.setItem(row, 0, date_item)

            # Тип
            type_item = QTableWidgetItem(data.get("type", ""))
            self.finance_table.setItem(row, 1, type_item)

            # Сумма
            amount_item = QTableWidgetItem(f"{data.get('amount', 0):,.2f} ₽".replace(",", " "))
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.finance_table.setItem(row, 2, amount_item)

            # Описание
            desc_item = QTableWidgetItem(data.get("description", ""))
            self.finance_table.setItem(row, 3, desc_item)

            # Детали
            details_item = QTableWidgetItem(data.get("details", ""))
            self.finance_table.setItem(row, 4, details_item)

    def update_finance_stats(self, finance_data):
        """Обновляет финансовую статистику"""
        self.revenue_stat.findChild(QLabel, "value").setText(
            f"{finance_data.get('revenue', 0):,.2f} ₽".replace(",", " "))
        self.expenses_stat.findChild(QLabel, "value").setText(
            f"{finance_data.get('expenses', 0):,.2f} ₽".replace(",", " "))
        self.profit_stat.findChild(QLabel, "value").setText(f"{finance_data.get('profit', 0):,.2f} ₽".replace(",", " "))
        self.orders_stat.findChild(QLabel, "value").setText(str(finance_data.get('orders_count', 0)))

    def update_finance_details(self, details):
        """Обновляет таблицу детализации"""
        self.finance_details_table.setRowCount(len(details))

        for row, detail in enumerate(details):
            # Дата
            date_str = detail.get("date", "")
            try:
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                date_item = QTableWidgetItem(date.strftime("%d.%m.%Y"))
            except ValueError:
                date_item = QTableWidgetItem(date_str)
            self.finance_details_table.setItem(row, 0, date_item)

            # Тип
            type_item = QTableWidgetItem(detail.get("type", ""))
            self.finance_details_table.setItem(row, 1, type_item)

            # Сумма
            amount = detail.get("amount", 0)
            amount_item = QTableWidgetItem(f"{amount:,.2f} ₽".replace(",", " "))
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

            # Цвет в зависимости от типа
            if detail.get("type") == "Доход":
                amount_item.setForeground(QColor("#2ecc71"))  # зеленый
            else:
                amount_item.setForeground(QColor("#e74c3c"))  # красный

            self.finance_details_table.setItem(row, 2, amount_item)

            # Описание
            desc_item = QTableWidgetItem(detail.get("description", ""))
            self.finance_details_table.setItem(row, 3, desc_item)

            # Детали
            details_text = ""
            if detail.get("order_id"):
                details_text = f"Заказ #{detail.get('order_id')}"
            elif detail.get("supply_id"):
                details_text = f"Поставка #{detail.get('supply_id')}"

            details_item = QTableWidgetItem(details_text)
            self.finance_details_table.setItem(row, 4, details_item)

    def generate_finance_report(self):
        """Генерирует финансовый отчет"""
        # Здесь должна быть логика генерации отчета (PDF, Excel и т.д.)
        QMessageBox.information(self, "Отчет", "Финансовый отчет сгенерирован и сохранен")

    def closeEvent(self, event):
        """Обработчик закрытия окна"""
        if QMessageBox.question(self, "Подтверждение",
                               "Вы действительно хотите выйти из системы?",
                               QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()